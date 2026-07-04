#!/usr/bin/env python3
"""
Excel 批量合并工具
把文件夹里所有 .xlsx 文件合并成一个
"""
import argparse
import glob
import os
import sys

try:
    import openpyxl
except ImportError:
    print("请先安装: pip install openpyxl")
    sys.exit(1)


def merge_excel(folder: str, output: str, sheet: str = None):
    """合并文件夹内所有 Excel"""
    files = glob.glob(os.path.join(folder, "*.xlsx"))
    if not files:
        print(f"❌ 在 {folder} 中没找到 .xlsx 文件")
        return

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "合并结果"

    header_written = False
    row = 1

    for f in sorted(files):
        print(f"📄 读取: {os.path.basename(f)}")
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb[sheet] if sheet else wb.active

        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                if not header_written:
                    ws_out.append(list(r) + ["来源文件"])
                    header_written = True
                continue
            ws_out.append(list(r) + [os.path.basename(f)])
            row += 1
        wb.close()

    wb_out.save(output)
    print(f"✅ 合并完成! {len(files)} 个文件 → {output} (共 {row} 行)")


def main():
    parser = argparse.ArgumentParser(description="Excel 批量合并")
    parser.add_argument("folder", help="包含 .xlsx 文件的文件夹")
    parser.add_argument("-o", "--output", default="合并结果.xlsx", help="输出文件名")
    parser.add_argument("-s", "--sheet", help="指定工作表名")
    args = parser.parse_args()

    merge_excel(args.folder, args.output, args.sheet)


if __name__ == "__main__":
    main()
